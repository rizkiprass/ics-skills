"""Generate XLSX summary from AWS Service Screener api-full.json with AWS scan enrichment."""
import json, re, sys, argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

PILLAR_MAP = {"R": "Reliability", "C": "Cost Optimization", "P": "Performance Efficiency", "O": "Operational Excellence", "S": "Security", "T": "Sustainability"}
SEVERITY_MAP = {"H": "High", "M": "Medium", "L": "Low", "I": "Informational"}
HEADERS = ["No", "Service", "Check", "Description", "Severity", "Region", "Affected Resources", "Resource Details", "Checklist", "Note"]

def strip_html(text):
    if not text: return ""
    return re.sub(r'<[^>]+>', '', str(text)).strip()

def clean_resource(r):
    return re.sub(r'^[A-Za-z0-9]+::', '', str(r)).replace('<b>', '').replace('</b>', '')

def get_resource_id(r):
    """Extract raw resource ID from screener format like 'SG::sg-xxx' or 'EC2::i-xxx'."""
    cleaned = re.sub(r'^[A-Za-z0-9]+::', '', str(r)).replace('<b>', '').replace('</b>', '')
    return cleaned.strip()

def enrich_sg(sg_id, scan_data):
    """Build detail string for a security group from scan data."""
    sgs = scan_data.get("security_groups", {})
    sg = sgs.get(sg_id)
    if not sg:
        return ""
    lines = [f"Name: {sg['name']}"]
    if sg.get("description"):
        lines.append(f"Desc: {sg['description']}")
    lines.append(f"VPC: {sg['vpc_id']}")
    if sg["ingress"]:
        lines.append("Inbound Rules:")
        for rule in sg["ingress"]:
            src = ", ".join(rule["sources"]) if rule["sources"] else "N/A"
            lines.append(f"  {rule['protocol']} port {rule['ports']} from {src}")
    return "\n".join(lines)

def enrich_ec2(instance_id, scan_data):
    """Build detail string for an EC2 instance from scan data."""
    instances = scan_data.get("ec2_instances", {})
    inst = instances.get(instance_id)
    if not inst:
        return ""
    lines = [f"Name: {inst['name']}", f"Type: {inst['type']}", f"State: {inst['state']}"]
    if inst.get("public_ip"):
        lines.append(f"Public IP: {inst['public_ip']}")
    if inst.get("private_ip"):
        lines.append(f"Private IP: {inst['private_ip']}")
    if inst.get("security_groups"):
        lines.append(f"SGs: {', '.join(inst['security_groups'])}")
    return "\n".join(lines)

def enrich_vpc(vpc_id, scan_data):
    """Build detail string for a VPC from scan data."""
    vpcs = scan_data.get("vpcs", {})
    vpc = vpcs.get(vpc_id)
    if not vpc:
        return ""
    lines = [f"Name: {vpc['name']}", f"CIDR: {vpc['cidr']}"]
    if vpc.get("is_default"):
        lines.append("(Default VPC)")
    return "\n".join(lines)

def enrich_elb(elb_name, scan_data):
    """Build detail string for an ELB from scan data."""
    elbs = scan_data.get("elbs", {})
    elb = elbs.get(elb_name)
    if not elb:
        return ""
    lines = [f"Type: {elb['type']}", f"Scheme: {elb['scheme']}"]
    if elb.get("security_groups"):
        lines.append(f"SGs: {', '.join(elb['security_groups'])}")
    if elb.get("listeners"):
        lines.append("Listeners:")
        for l in elb["listeners"]:
            ssl = f" (SSL: {l['ssl_policy']})" if l.get("ssl_policy") else ""
            lines.append(f"  {l['protocol']}:{l['port']}{ssl}")
    return "\n".join(lines)

def enrich_iam_user(username, scan_data):
    """Build detail string for an IAM user from scan data."""
    users = scan_data.get("iam_users", {})
    user = users.get(username)
    if not user:
        return ""
    lines = [f"MFA: {'Enabled' if user['mfa_enabled'] else 'NOT Enabled'}"]
    if user.get("groups"):
        lines.append(f"Groups: {', '.join(user['groups'])}")
    if user.get("attached_policies"):
        lines.append(f"Policies: {', '.join(user['attached_policies'])}")
    return "\n".join(lines)

def enrich_resource(resource_raw, scan_data):
    """Determine resource type and enrich with scan data."""
    rid = get_resource_id(resource_raw)
    if rid.startswith("sg-"):
        return enrich_sg(rid, scan_data)
    elif rid.startswith("i-"):
        return enrich_ec2(rid, scan_data)
    elif rid.startswith("vpc-"):
        return enrich_vpc(rid, scan_data)
    # Check IAM users
    users = scan_data.get("iam_users", {})
    if rid in users:
        return enrich_iam_user(rid, scan_data)
    # Check ELBs
    elbs = scan_data.get("elbs", {})
    if rid in elbs:
        return enrich_elb(rid, scan_data)
    return ""

def resolve_resource_name(resource_raw, scan_data):
    """Resolve a resource ID to a friendly 'Name (id)' string using scan data."""
    if not scan_data:
        return clean_resource(resource_raw)
    rid = get_resource_id(resource_raw)
    # EC2 instances
    if rid.startswith("i-"):
        inst = scan_data.get("ec2_instances", {}).get(rid)
        if inst and inst.get("name"):
            return f"{inst['name']} ({rid})"
    # Security Groups
    elif rid.startswith("sg-"):
        sg = scan_data.get("security_groups", {}).get(rid)
        if sg and sg.get("name"):
            return f"{sg['name']} ({rid})"
    # VPCs
    elif rid.startswith("vpc-"):
        vpc = scan_data.get("vpcs", {}).get(rid)
        if vpc and vpc.get("name"):
            return f"{vpc['name']} ({rid})"
    return clean_resource(resource_raw)

def enrich_resources_list(resources_raw, scan_data):
    """Enrich a list of raw resources and combine details."""
    if not scan_data:
        return ""
    details = []
    for r in resources_raw:
        detail = enrich_resource(r, scan_data)
        if detail:
            rid = get_resource_id(r)
            details.append(f"[{rid}]\n{detail}")
    return "\n\n".join(details)

def parse_findings(data, severities, scan_data=None):
    findings = {}
    for service, svc_data in data.items():
        summary = svc_data.get("summary", {})
        for check_name, check in summary.items():
            crit = check.get("criticality", "")
            if crit not in severities: continue
            pillar = check.get("__categoryMain", "")
            if pillar not in PILLAR_MAP: continue
            affected = check.get("__affectedResources", {})
            for region, resources in affected.items():
                resource_details = enrich_resources_list(resources, scan_data) if scan_data else ""
                entry = {
                    "service": service,
                    "check": check_name,
                    "shortDesc": strip_html(check.get("shortDesc", "")),
                    "description": strip_html(check.get("^description", check.get("shortDesc", ""))),
                    "severity": SEVERITY_MAP.get(crit, crit),
                    "region": region,
                    "resources": ", ".join(resolve_resource_name(r, scan_data) for r in resources),
                    "resource_details": resource_details,
                    "checklist": "",
                    "note": ""
                }
                findings.setdefault(pillar, []).append(entry)
    return findings

def generate_md(findings, output_path):
    lines = ["# AWS Service Screener Summary\n"]
    for code in ["R", "C", "P", "O", "S", "T"]:
        items = findings.get(code, [])
        if not items: continue
        lines.append(f"\n## {PILLAR_MAP[code]}\n")
        lines.append("| No | Service | Check | Description | Severity | Region | Affected Resources | Resource Details | Checklist | Note |")
        lines.append("|----|---------|-------|-------------|----------|--------|--------------------|------------------|-----------|------|")
        for i, f in enumerate(items, 1):
            details = f['resource_details'].replace('\n', ' | ') if f['resource_details'] else "-"
            lines.append(f"| {i} | {f['service']} | {f['shortDesc']} | {f['description']} | {f['severity']} | {f['region']} | {f['resources']} | {details} | [ ] | |")
    Path(output_path).write_text("\n".join(lines), encoding="utf-8")
    print(f"Markdown saved: {output_path}")

def generate_xlsx(findings, output_path):
    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    cell_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    col_widths = [5, 14, 25, 55, 10, 16, 40, 50, 12, 30]
    checklist_dv = DataValidation(type="list", formula1='"Done,In Progress,Not Started,N/A"', allow_blank=True)
    checklist_dv.error = "Please select a valid status"
    checklist_dv.errorTitle = "Invalid Status"

    for code in ["R", "C", "P", "O", "S", "T"]:
        items = findings.get(code, [])
        if not items: continue
        ws = wb.create_sheet(title=PILLAR_MAP[code][:31])
        ws.add_data_validation(checklist_dv)
        for col_idx, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for row_idx, f in enumerate(items, 2):
            vals = [row_idx - 1, f["service"], f["shortDesc"], f["description"],
                    f["severity"], f["region"], f["resources"], f["resource_details"],
                    "Not Started", f["note"]]
            for col_idx, v in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=v)
                cell.alignment = cell_align
                cell.border = thin_border
                cell.font = cell_font
            checklist_dv.add(ws.cell(row=row_idx, column=9))
        for col_idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    if not wb.sheetnames:
        ws = wb.create_sheet("No Findings")
        ws["A1"] = "No findings matched the filter criteria."

    wb.save(output_path)
    print(f"XLSX saved: {output_path}")

def main():
    parser = argparse.ArgumentParser(description="AWS Service Screener Summary Generator")
    parser.add_argument("input", help="Path to api-full.json")
    parser.add_argument("output", help="Output file path (.xlsx)")
    parser.add_argument("--severity", default="H", help="Severity filter, comma-separated (H,M,L,I). Default: H")
    parser.add_argument("--md", help="Also generate markdown file at this path")
    parser.add_argument("--scan", help="Path to AWS scan results JSON (from aws_scanner.py)")
    parser.add_argument("--live-scan", action="store_true", help="Run live AWS scan before generating report")
    parser.add_argument("--region", help="AWS region for live scan")
    parser.add_argument("--env", default=".env", help="Path to .env file for AWS credentials")
    args = parser.parse_args()

    severities = [s.strip() for s in args.severity.split(",")]
    with open(args.input, "r", encoding="utf-8") as f:
        data = json.load(f)

    scan_data = None
    if args.live_scan:
        from aws_scanner import load_env, scan_all
        if args.env:
            load_env(args.env)
        scan_data = scan_all(args.region)
        scan_output = Path(args.output).with_suffix(".scan.json")
        scan_output.write_text(json.dumps(scan_data, indent=2, default=str), encoding="utf-8")
        print(f"Scan results saved: {scan_output}")
    elif args.scan:
        with open(args.scan, "r", encoding="utf-8") as f:
            scan_data = json.load(f)

    findings = parse_findings(data, severities, scan_data)
    if args.md:
        generate_md(findings, args.md)
    generate_xlsx(findings, args.output)

if __name__ == "__main__":
    main()
